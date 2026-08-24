#!/usr/bin/env python3
"""Inspect and validate the monthly roster workbooks stored in this repository.

This repository is a data store: each ``*.xlsx`` file is an operations roster
(employee names plus a day-by-day shift grid). This utility opens the workbooks
with openpyxl so you can quickly confirm they are readable and see a summary of
their structure without launching a spreadsheet application.

Usage:
    python scripts/inspect_roster.py                 # summary of every workbook
    python scripts/inspect_roster.py --validate      # exit non-zero if any file fails to open
    python scripts/inspect_roster.py "IMP FEB 2026.xlsx" --preview
"""
from __future__ import annotations

import argparse
import glob
import os
import sys

import openpyxl

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def find_workbooks() -> list[str]:
    """Return repository ``*.xlsx`` files sorted by name."""
    pattern = os.path.join(REPO_ROOT, "*.xlsx")
    return sorted(glob.glob(pattern))


def _first_text_cell(worksheet, max_scan: int = 12) -> str:
    """Return the first non-empty text cell, used as a human-friendly title."""
    for row in worksheet.iter_rows(min_row=1, max_row=max_scan, values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                return cell.strip()
    return ""


def summarize(path: str) -> dict:
    """Open one workbook and collect a small structural summary."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for name in workbook.sheetnames:
            worksheet = workbook[name]
            sheets.append(
                {
                    "name": name,
                    "rows": worksheet.max_row,
                    "cols": worksheet.max_column,
                    "title": _first_text_cell(worksheet),
                }
            )
        return {"sheets": sheets}
    finally:
        workbook.close()


def preview(path: str, rows: int = 6, cols: int = 8) -> None:
    """Print the top-left corner of the first sheet."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        print(f"  preview of sheet '{worksheet.title}':")
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            cells = ["" if c is None else str(c)[:14] for c in row[:cols]]
            print("    " + " | ".join(cells))
            if index + 1 >= rows:
                break
    finally:
        workbook.close()


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "files",
        nargs="*",
        help="Specific workbook file(s) to inspect (defaults to every *.xlsx in the repo).",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="Exit with a non-zero status if any workbook fails to open.",
    )
    parser.add_argument(
        "--preview",
        action="store_true",
        help="Print the first few rows of each workbook's first sheet.",
    )
    args = parser.parse_args(argv)

    targets = args.files or [os.path.relpath(p, REPO_ROOT) for p in find_workbooks()]
    if not targets:
        print("No .xlsx roster workbooks found in the repository.", file=sys.stderr)
        return 1

    failures = 0
    for name in targets:
        path = name if os.path.isabs(name) else os.path.join(REPO_ROOT, name)
        if not os.path.exists(path):
            print(f"[MISSING] {name}", file=sys.stderr)
            failures += 1
            continue
        try:
            info = summarize(path)
        except Exception as exc:  # noqa: BLE001 - report any parse failure
            print(f"[FAIL] {name}: {exc}", file=sys.stderr)
            failures += 1
            continue

        sheet_desc = ", ".join(
            f"{s['name']} ({s['rows']}x{s['cols']})" for s in info["sheets"]
        )
        print(f"[OK] {name}")
        print(f"     sheets: {sheet_desc}")
        if args.preview:
            preview(path)

    total = len(targets)
    print(f"\n{total - failures}/{total} workbook(s) opened successfully.")
    if args.validate and failures:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
